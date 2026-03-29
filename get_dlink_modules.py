import time
import re
import datetime
import pandas as pd
from netmiko import ConnectHandler
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.utils import get_column_letter
# 👇 【新規追加】プルダウンと条件付き書式のためのライブラリ
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

HOSTS = [
    {"hostname": "ES-KFK-OFC-01", "ip": "172.19.151.254"},
    {"hostname": "ES-WEL1-OFC-01", "ip": "172.19.161.254"},
    {"hostname": "ES-STA1-EPS-01", "ip": "172.19.171.254"},
    {"hostname": "ES-EST3-SVR-01", "ip": "172.19.181.254"},
    {"hostname": "BS-1BLB1-EPS-01", "ip": "172.19.11.254"},
    {"hostname": "BS-2BL3-EPS-01", "ip": "172.19.21.254"},
    {"hostname": "BS-OGDB1-WRH-01", "ip": "172.19.81.254"},
    {"hostname": "BS-LIB1-PWR-01", "ip": "172.19.91.254"},
    {"hostname": "BS-KRBB-SVR01", "ip": "172.19.1.254"}
]

OFFICIAL_MODULES = [
    "DEM-431XT", "DEM-432XT", "DEM-410T", # 10G
    "DEM-311GT", "DEM-310GT", "DEM-330T", "DEM-330R", "DGS-712", # 1G
    "DEM-211" # 100M
]

MODULE_STANDARDS = {
    "DEM-431XT": "10GBASE-SR",
    "DEM-432XT": "10GBASE-LR",
    "DEM-410T": "10GBASE-T",
    "DEM-311GT": "1000BASE-SX",
    "DEM-310GT": "1000BASE-LX",
    "DEM-330T": "1000BASE-BX-D",
    "DEM-330R": "1000BASE-BX-U",
    "DGS-712": "1000BASE-T",
    "DEM-211": "100BASE-FX",
    "コンボポート": "1000BASE-T",
    "不明・非純正 (reading...)": "-",
    "空きスロット (未接続/ダウン)": "-"
}

def get_password(hostname):
    if hostname.startswith("ES"):
        return "tfuES916!"
    elif hostname.startswith("BS"):
        return "tfuBS916!"
    return ""

def determine_connection_type(port_type, vendor_pn, compliance, raw_gbic_output):
    if "1000BASE-T" in port_type or "1000BASE-T" in compliance or vendor_pn in ["DEM-410T", "DGS-712"]:
        return "銅 (Copper)"
    if "reading" in vendor_pn.lower():
        return "不明 (reading)"
    
    gbic_upper = raw_gbic_output.upper()
    if "MULTI-MODE" in gbic_upper or "OM1" in gbic_upper or "OM2" in gbic_upper or "OM3" in gbic_upper:
        return "マルチ光 (MMF)"
    if "SINGLE-MODE" in gbic_upper or "SINGLE MODE" in gbic_upper or "9/125" in gbic_upper:
        return "シングル光 (SMF)"

    comp_upper = compliance.upper()
    if "SX" in comp_upper or "SR" in comp_upper or vendor_pn in ["DEM-311GT", "DEM-431XT", "DEM-211"]:
        return "マルチ光 (MMF)"
    if "LX" in comp_upper or "LR" in comp_upper or "BX" in comp_upper or vendor_pn in ["DEM-310GT", "DEM-432XT", "DEM-330T", "DEM-330R"]:
        return "シングル光 (SMF)"
    
    if "LC" in port_type or "SFP" in port_type:
        return "光 (詳細不明)"
    
    return "-"

def collect_switch_data(host_info):
    hostname = host_info["hostname"]
    ip = host_info["ip"]
    password = get_password(hostname)
    
    print(f"\n[{hostname}] ({ip}) への接続を開始します...")
    
    log_filename = f"{hostname}_teraterm_log.txt"
    device = {
        'device_type': 'cisco_ios', 
        'host': ip,
        'username': 'admin',
        'password': password,
        'global_delay_factor': 2,
        'session_log': log_filename
    }
    
    port_data = []
    
    try:
        with ConnectHandler(**device) as net_connect:
            net_connect.send_command("terminal length 0")
            
            desc_output = net_connect.send_command("show interfaces description")
            descriptions = {}
            for line in desc_output.splitlines():
                parts = line.split()
                if len(parts) >= 2 and parts[0].startswith("eth"):
                    desc_text = " ".join(parts[3:]) if len(parts) > 3 else ""
                    descriptions[parts[0]] = desc_text.strip(" *")

            status_output = net_connect.send_command("show interfaces status")
            
            for line in status_output.splitlines():
                parts = line.split()
                if not parts or not parts[0].startswith("eth"):
                    continue
                
                port = parts[0]
                status = parts[1] if len(parts) > 1 else ""
                duplex = parts[3] if len(parts) > 3 else ""
                speed = parts[4] if len(parts) > 4 else ""
                port_type = parts[5] if len(parts) > 5 else ""
                desc = descriptions.get(port, "")
                
                vendor_pn = "-"
                compliance = "-"
                raw_gbic_output = ""
                
                port_id_for_cmd = port.replace("eth", "").replace("(c)", "").replace("(f)", "")
                
                if status == "connected" and port_type != "1000BASE-T" and "(c)" not in port:
                    gbic_cmd = f"show interfaces ethernet {port_id_for_cmd} gbic"
                    gbic_output = net_connect.send_command(gbic_cmd)
                    
                    if "reading..." in gbic_output:
                        time.sleep(1)
                        gbic_output = net_connect.send_command(gbic_cmd)
                    
                    raw_gbic_output = gbic_output
                    
                    if "reading..." in gbic_output:
                        vendor_pn = "不明 (reading...)"
                    else:
                        pn_match = re.search(r"Vendor PN:\s*(.+)", gbic_output)
                        comp_match = re.search(r"Ethernet Compliance Code:\s*(.+)", gbic_output)
                        
                        if pn_match:
                            vendor_pn = pn_match.group(1).strip()
                        if comp_match:
                            compliance = comp_match.group(1).strip()
                            if compliance == "-":
                                compliance = ""
                                
                        if not compliance and vendor_pn in MODULE_STANDARDS:
                            compliance = MODULE_STANDARDS[vendor_pn]
                
                elif status != "connected":
                    vendor_pn = "未接続"
                elif "(c)" in port or port_type == "1000BASE-T":
                    vendor_pn = "(RJ-45 内蔵)"
                    compliance = "1000BASE-T"

                conn_type = determine_connection_type(port_type, vendor_pn, compliance, raw_gbic_output)
                
                port_data.append({
                    "ポート番号": port,
                    "リンク状態": status,
                    "接続方式": conn_type,
                    "モジュール型番": vendor_pn,
                    "規格": compliance,
                    "速度 (Speed)": speed,
                    "Duplex": duplex,
                    "Type": port_type,
                    "接続先 (Description)": desc,
                    "更改後の規格": compliance 
                })
                
        print(f"[{hostname}] データの取得が完了しました。")
        return hostname, port_data

    except Exception as e:
        print(f"[{hostname}] エラーが発生しました: {e}")
        return hostname, []

def main():
    all_data = {}
    summary_data = []

    for host in HOSTS:
        hostname, data = collect_switch_data(host)
        if data:
            all_data[hostname] = data
            
            counts = {mod: 0 for mod in OFFICIAL_MODULES}
            counts["コンボポート"] = 0
            counts["不明・非純正 (reading...)"] = 0
            counts["空きスロット (未接続/ダウン)"] = 0
            
            for row in data:
                pn = row["モジュール型番"]
                status = row["リンク状態"]
                
                if status != "connected":
                    counts["空きスロット (未接続/ダウン)"] += 1
                elif "reading" in pn:
                    counts["不明・非純正 (reading...)"] += 1
                elif "RJ-45" in pn or row["接続方式"] == "銅 (Copper)":
                    counts["コンボポート"] += 1
                else:
                    matched = False
                    for mod in OFFICIAL_MODULES:
                        if mod in pn:
                            counts[mod] += 1
                            matched = True
                            break
                    if not matched:
                        counts[pn] = counts.get(pn, 0) + 1
            
            summary_row = {"ホスト名": hostname}
            summary_row.update(counts)
            summary_data.append(summary_row)

    if not all_data:
        print("\nデータが一つも取得できませんでした。")
        return

    now_time = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = f"DGS_Module_Inventory_Complete_{now_time}.xlsx"
    
    # 🎨 スタイル定義
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_font = Font(bold=True)
    
    # 🌟 モジュール規格別の色
    fill_10g_sr = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid") # 黄色
    fill_10g_lr = PatternFill(start_color="9BC2E6", end_color="9BC2E6", fill_type="solid") # 水色（新規: 10GBASE-LR）
    fill_1g_sx  = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # 薄い黄色
    fill_100m_fx = PatternFill(start_color="E4DFEC", end_color="E4DFEC", fill_type="solid") # 薄い紫
    fill_1g_t   = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid") # 薄い橙
    
    fill_connected = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") # 薄い緑

    # 📝 レジェンドのデータ
    legend_data = [
        ("10GBASE-SRは速いから黄色", fill_10g_sr),
        ("10GBASE-LRは長距離だから水色(更改後に追加)", fill_10g_lr),
        ("1000BASE-SXはそこそこだし光だから薄い黄色", fill_1g_sx),
        ("100BASE-FXは問題児だから薄い紫", fill_100m_fx),
        ("1000BASE-Tは銅だから薄いだいたい色", fill_1g_t)
    ]

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        
        # --- シート1（総合集計） ---
        df_summary = pd.DataFrame(summary_data)
        cols_to_keep = ["ホスト名"]
        for col in df_summary.columns:
            if col == "ホスト名": continue
            if (df_summary[col] != 0).any():
                cols_to_keep.append(col)
        df_summary = df_summary[cols_to_keep]

        special_order = ["コンボポート", "不明・非純正 (reading...)", "空きスロット (未接続/ダウン)"]
        sorted_cols = ["ホスト名"]
        for mod in OFFICIAL_MODULES:
            if mod in df_summary.columns:
                sorted_cols.append(mod)
        for col in df_summary.columns:
            if col not in sorted_cols and col not in special_order:
                sorted_cols.append(col)
        for spec in special_order:
            if spec in df_summary.columns:
                sorted_cols.append(spec)
        df_summary = df_summary[sorted_cols]

        df_summary.to_excel(writer, sheet_name="00_総合集計", index=False, header=False, startrow=2)
        
        ws_summary = writer.sheets["00_総合集計"]
        ws_summary.freeze_panes = "B3"
        
        max_row = len(df_summary) + 2
        for col_idx, col_name in enumerate(df_summary.columns, start=1):
            cell_1 = ws_summary.cell(row=1, column=col_idx)
            cell_2 = ws_summary.cell(row=2, column=col_idx)
            
            if col_name == "ホスト名":
                cell_1.value = "ホスト名"
                ws_summary.merge_cells(start_row=1, start_column=col_idx, end_row=2, end_column=col_idx)
            else:
                standard = MODULE_STANDARDS.get(col_name, "規格不明")
                cell_1.value = col_name
                cell_2.value = standard
                
                col_fill = None
                if standard == "10GBASE-SR": col_fill = fill_10g_sr
                elif standard == "1000BASE-SX": col_fill = fill_1g_sx
                elif standard == "100BASE-FX": col_fill = fill_100m_fx
                elif standard == "1000BASE-T": col_fill = fill_1g_t
                
                if col_fill:
                    for r in range(3, max_row + 1):
                        ws_summary.cell(row=r, column=col_idx).fill = col_fill
                
            for cell in [cell_1, cell_2]:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = header_fill
                cell.font = header_font

        # --- シート2以降（各ホスト詳細） ---
        for hostname, data in all_data.items():
            safe_sheet_name = hostname[:31] 
            df_details = pd.DataFrame(data)
            df_details.to_excel(writer, sheet_name=safe_sheet_name, index=False)
            
            ws_detail = writer.sheets[safe_sheet_name]
            ws_detail.auto_filter.ref = ws_detail.dimensions
            ws_detail.freeze_panes = "A2"
            
            max_r = len(df_details) + 1
            
            # 📌 【新規】J列（更改後の規格）のプルダウン（入力規則）設定
            dv = DataValidation(type="list", formula1='"10GBASE-SR,10GBASE-LR,1000BASE-SX,100BASE-FX,1000BASE-T,-"', allow_blank=True)
            dv.error = '無効な規格です。リストから選択してください。'
            dv.errorTitle = '入力エラー'
            ws_detail.add_data_validation(dv)
            dv.add(f"J2:J{max_r}")
            
            # 📌 【新規】J列（更改後の規格）の条件付き書式（自動色付け）
            j_range = f"J2:J{max_r}"
            ws_detail.conditional_formatting.add(j_range, CellIsRule(operator='equal', formula=['"10GBASE-SR"'], stopIfTrue=True, fill=fill_10g_sr))
            ws_detail.conditional_formatting.add(j_range, CellIsRule(operator='equal', formula=['"10GBASE-LR"'], stopIfTrue=True, fill=fill_10g_lr))
            ws_detail.conditional_formatting.add(j_range, CellIsRule(operator='equal', formula=['"1000BASE-SX"'], stopIfTrue=True, fill=fill_1g_sx))
            ws_detail.conditional_formatting.add(j_range, CellIsRule(operator='equal', formula=['"100BASE-FX"'], stopIfTrue=True, fill=fill_100m_fx))
            ws_detail.conditional_formatting.add(j_range, CellIsRule(operator='equal', formula=['"1000BASE-T"'], stopIfTrue=True, fill=fill_1g_t))

            # ヘッダー装飾
            for c_idx in range(1, len(df_details.columns) + 1):
                cell = ws_detail.cell(row=1, column=c_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # 🎨 データ行の色塗り（disabled等には色を塗らない）
            for r_idx in range(2, max_r + 1):
                status_val = df_details.iloc[r_idx-2]["リンク状態"]
                standard_val = df_details.iloc[r_idx-2]["規格"]
                
                row_fill = None
                
                if status_val == "connected":
                    if standard_val == "10GBASE-SR": row_fill = fill_10g_sr
                    elif standard_val == "1000BASE-SX": row_fill = fill_1g_sx
                    elif standard_val == "100BASE-FX": row_fill = fill_100m_fx
                    elif standard_val == "1000BASE-T": row_fill = fill_1g_t
                    else:
                        row_fill = fill_connected
                        
                if row_fill:
                    # J列（更改後）も初期値として同じ色に塗られますが、値を変更すると条件付き書式が優先されて色が変わります
                    for c_idx in range(1, len(df_details.columns) + 1):
                        ws_detail.cell(row=r_idx, column=c_idx).fill = row_fill

            # 📌 レジェンドの挿入 (K列 2行目〜)
            legend_col = 11
            for i, (text, fill_color) in enumerate(legend_data):
                r = 2 + i
                ws_detail.cell(row=r, column=legend_col).value = text
                ws_detail.cell(row=r, column=legend_col).fill = fill_color
                ws_detail.cell(row=r, column=legend_col).alignment = Alignment(vertical='center')

            # 📌 個別集計データの挿入 (K列 8行目・9行目)
            host_summary_row = df_summary[df_summary["ホスト名"] == hostname]
            if not host_summary_row.empty:
                summary_cols = list(df_summary.columns)
                summary_vals = host_summary_row.iloc[0].tolist()

                for i, col_name in enumerate(summary_cols):
                    h_cell = ws_detail.cell(row=8, column=legend_col + i)
                    h_cell.value = col_name
                    h_cell.fill = header_fill
                    h_cell.font = header_font
                    h_cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    d_cell = ws_detail.cell(row=9, column=legend_col + i)
                    d_cell.value = summary_vals[i]
                    d_cell.alignment = Alignment(horizontal='center', vertical='center')

            # 📌 J列のカウント機能 (K列 11行目以降)
            count_start_row = 11
            title_cell = ws_detail.cell(row=count_start_row, column=legend_col)
            title_cell.value = "▼ 【更改後の規格】 J列の文字をカウント"
            title_cell.font = Font(bold=True)
            
            count_items = [
                ("10GBASE-SR", fill_10g_sr),
                ("10GBASE-LR", fill_10g_lr),
                ("1000BASE-SX", fill_1g_sx),
                ("100BASE-FX", fill_100m_fx),
                ("1000BASE-T", fill_1g_t)
            ]
            
            for i, (spec, fill_c) in enumerate(count_items):
                r = count_start_row + 1 + i
                
                # 規格名
                spec_cell = ws_detail.cell(row=r, column=legend_col)
                spec_cell.value = spec
                spec_cell.fill = fill_c
                spec_cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # COUNTIF関数
                formula_cell = ws_detail.cell(row=r, column=legend_col + 1)
                formula_cell.value = f'=COUNTIF(J:J, "{spec}")'
                formula_cell.font = Font(bold=True)
                formula_cell.alignment = Alignment(horizontal='center', vertical='center')

            # 列幅の自動調整
            max_col_to_adjust = max(len(df_details.columns), legend_col + len(df_summary.columns) - 1)
            for col_idx in range(1, max_col_to_adjust + 1):
                column_letter = get_column_letter(col_idx)
                max_length = 0
                for r_idx in range(1, ws_detail.max_row + 1):
                    cell_val = ws_detail.cell(row=r_idx, column=col_idx).value
                    if cell_val:
                        try:
                            if str(cell_val).startswith('='): continue
                            val_len = len(str(cell_val).encode('shift_jis', errors='ignore'))
                            if val_len > max_length:
                                max_length = val_len
                        except:
                            pass
                ws_detail.column_dimensions[column_letter].width = min(max_length + 2, 60)

        # 総合集計シートの下部レジェンド
        legend_start_row = ws_summary.max_row + 2
        for i, (text, fill_color) in enumerate(legend_data):
            r = legend_start_row + i
            ws_summary.cell(row=r, column=2).value = text
            ws_summary.cell(row=r, column=2).fill = fill_color
            ws_summary.cell(row=r, column=2).alignment = Alignment(vertical='center')

    print(f"\n✨ 全完了！ '{output_file}' に結果を出力しました。")

if __name__ == "__main__":
    main()